"""
Извлечение текста из PDF, Excel, JSON для последующего чанкинга.
"""
from __future__ import annotations

import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)


def _pdf_reader_class():
    """pypdf (рекомендуется) или PyPDF2 из requirements."""
    try:
        from pypdf import PdfReader as PR

        return PR, "pypdf"
    except ImportError:
        pass
    try:
        from PyPDF2 import PdfReader as PR

        return PR, "PyPDF2"
    except ImportError:
        pass
    raise ImportError(
        "Для PDF нужен пакет pypdf. В каталоге backend выполните: pip install pypdf"
    )


def parse_pdf(data: bytes) -> str:
    """Текст из PDF (постранично, абзацы через перенос)."""
    PdfReader, _lib = _pdf_reader_class()
    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    for page in reader.pages:
        try:
            t = page.extract_text() or ""
        except Exception as e:
            logger.warning("PDF страница: ошибка извлечения: %s", e)
            t = ""
        t = t.strip()
        if t:
            parts.append(t)
    return "\n\n".join(parts)


def parse_xlsx(data: bytes) -> list[tuple[str, str]]:
    """
    Каждый элемент: (имя_листа, текстовый блок).
    Строка таблицы = «колонка: значение» через точку с запятой; блоки по группам строк.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "Для Excel нужен openpyxl: pip install openpyxl"
        ) from e
    import os

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows_per_block = int(os.getenv("RAG_XLSX_ROWS_PER_BLOCK", "12"))
    out: list[tuple[str, str]] = []

    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
        buf: list[str] = []
        row_num = 1
        for row in rows_iter:
            row_num += 1
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            pairs = []
            for i, h in enumerate(headers):
                if i >= len(row):
                    break
                v = row[i]
                if v is None or str(v).strip() == "":
                    continue
                pairs.append(f"{h}: {v}")
            if pairs:
                buf.append(f"Строка {row_num}: " + "; ".join(pairs))
            if len(buf) >= rows_per_block:
                out.append((sheet.title, "\n".join(buf)))
                buf = []
        if buf:
            out.append((sheet.title, "\n".join(buf)))
    wb.close()
    return out


def _json_to_text_chunks(obj: Any, path: str = "$") -> list[tuple[str, str]]:
    """Пары (json_path_hint, текст для эмбеддинга)."""
    chunks: list[tuple[str, str]] = []
    if isinstance(obj, dict):
        if not obj:
            chunks.append((path, "{}"))
            return chunks
        for k, v in obj.items():
            p = f"{path}.{k}" if path != "$" else f"$.{k}"
            if isinstance(v, (dict, list)):
                chunks.extend(_json_to_text_chunks(v, p))
            else:
                chunks.append((p, f"{k}: {json.dumps(v, ensure_ascii=False)}"))
    elif isinstance(obj, list):
        if not obj:
            chunks.append((path, "[]"))
            return chunks
        for i, item in enumerate(obj):
            p = f"{path}[{i}]"
            if isinstance(item, (dict, list)):
                chunks.extend(_json_to_text_chunks(item, p))
            else:
                chunks.append((p, json.dumps(item, ensure_ascii=False)))
    else:
        chunks.append((path, json.dumps(obj, ensure_ascii=False)))
    return chunks


def parse_json_bytes(data: bytes) -> list[tuple[str, str]]:
    """JSON → список смысловых фрагментов с путём в метаданные."""
    raw = data.decode("utf-8-sig")
    obj = json.loads(raw)
    return json_obj_to_chunks(obj)


def json_obj_to_chunks(obj: Any) -> list[tuple[str, str]]:
    """Объект Python после json.loads — те же пары (путь, текст)."""
    return _json_to_text_chunks(obj, "$")


def parse_docx(data: bytes) -> str:
    """Текст из .docx (Word) — параграфы через двойной перенос."""
    try:
        from docx import Document
    except ImportError as e:
        raise ImportError(
            "Для Word (.docx) нужен python-docx: pip install python-docx"
        ) from e
    doc = Document(io.BytesIO(data))
    parts: list[str] = []
    for para in doc.paragraphs:
        t = para.text.strip()
        if t:
            parts.append(t)
    return "\n\n".join(parts)


def parse_csv(data: bytes) -> list[tuple[str, str]]:
    """CSV → список блоков строк (аналогично Excel)."""
    import csv
    import os

    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_per_block = int(os.getenv("RAG_XLSX_ROWS_PER_BLOCK", "12"))
    try:
        headers = next(reader)
    except StopIteration:
        return []
    headers = [h.strip() or f"col{i}" for i, h in enumerate(headers)]
    buf: list[str] = []
    row_num = 1
    out: list[tuple[str, str]] = []
    for row in reader:
        row_num += 1
        if not any(v.strip() for v in row):
            continue
        pairs = [f"{headers[i]}: {v.strip()}" for i, v in enumerate(row) if i < len(headers) and v.strip()]
        if pairs:
            buf.append(f"Строка {row_num}: " + "; ".join(pairs))
        if len(buf) >= rows_per_block:
            out.append(("CSV", "\n".join(buf)))
            buf = []
    if buf:
        out.append(("CSV", "\n".join(buf)))
    return out
